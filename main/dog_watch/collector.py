"""
Dog Watch data collector — runs independently of the web app.

Entry point:  python manage.py dog_watch_sync

Configure on Render as a Cron Job (not the web service):
  Schedule:  0 6 * * *   (daily at 6 AM UTC)
  Command:   python manage.py dog_watch_sync

Each run refreshes the USDA licensee list first, then checks APHIS for new reports.
"""
import io
import logging
import time
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeout
from datetime import date, datetime, timedelta

import openpyxl
import requests
from django.db.models import Q
from django.utils import timezone

from main.models import PuppyMillFacility
from main.dog_watch.aphis_client import search_inspections
from main.dog_watch.geocoder import geocode, normalize_state
from main.dog_watch.report_address import fetch_address_from_report_url
from main.dog_watch.report_sync import (
    create_new_inspection_reports,
    find_new_inspections,
    parse_inspection_date,
    parse_pending_violations_batch,
    parse_violations_for_reports,
    recompute_facility_violation_counts,
)
from main.dog_watch import sync_state

logger = logging.getLogger(__name__)

USDA_LICENSEE_URL = (
    'https://www.aphis.usda.gov/sites/default/files/'
    'list-of-active-licensees-and-registrants.xlsx'
)
TARGET_LICENSE_TYPES = {'Class A - Breeder', 'Class B - Dealer'}
HEADER_ROW_LABEL = 'Registration Type'


def _parse_expiration(value) -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    try:
        return datetime.strptime(str(value).strip(), '%m/%d/%Y').date()
    except ValueError:
        return None


def fetch_usda_licensees() -> list[dict]:
    """Download and parse the USDA active licensees spreadsheet."""
    res = requests.get(USDA_LICENSEE_URL, timeout=60)
    res.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(res.content), read_only=True)
    ws = wb[wb.sheetnames[0]]

    facilities = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        if row[1] == HEADER_ROW_LABEL:
            headers = row
            continue
        if not headers or not row[1] or not str(row[1]).startswith('Class '):
            continue
        license_type = row[1]
        if license_type not in TARGET_LICENSE_TYPES:
            continue
        row_data = dict(zip(headers, row))
        license_number = row_data.get('APHIS Registration Number') or row_data.get('APHIS License Number') or ''
        if not license_number:
            continue
        facilities.append({
            'license_number': str(license_number).strip(),
            'license_type': license_type,
            'name': (row_data.get('Account Name') or '').strip(),
            'dba_name': (row_data.get('DBA Name(s)') or '').strip(),
            'city': (row_data.get('Mailing City') or row_data.get('City') or '').strip(),
            'state': (row_data.get('State Abbreviation') or row_data.get('State') or '').strip(),
            'license_expiration': _parse_expiration(row_data.get('Expiration Date')),
            'usda_profile_url': (
                f'https://efile.aphis.usda.gov/PublicSearchTool/s/'
                f'inspection-reports?certNumber={license_number}'
            ),
        })
    return facilities


def _geocode_facility(facility: PuppyMillFacility) -> bool:
    state = normalize_state(facility.state)
    if not facility.city or not state:
        return False
    lat, lng, geocoded = geocode(
        facility.name, facility.city, state, facility.street_address, facility.zip_code,
    )
    if not geocoded or lat is None or lng is None:
        return False
    facility.latitude = lat
    facility.longitude = lng
    facility.coordinates_geocoded = True
    if state != facility.state:
        facility.state = state
    return True


def _newest_inspection_url(inspections: list[dict]) -> str:
    if not inspections:
        return ''
    dated = [
        insp for insp in inspections
        if insp.get('reportLink')
        and parse_inspection_date(insp.get('inspectionDate') or insp.get('inspectionDateString'))
    ]
    if dated:
        return max(
            dated,
            key=lambda i: parse_inspection_date(
                i.get('inspectionDate') or i.get('inspectionDateString')
            ) or date.min,
        ).get('reportLink') or ''
    return inspections[0].get('reportLink') or ''


def _apply_pdf_address(facility: PuppyMillFacility, pdf_address: dict) -> bool:
    changed = False
    if pdf_address.get('street_address') and pdf_address['street_address'] != facility.street_address:
        facility.street_address = pdf_address['street_address']
        changed = True
    if pdf_address.get('city') and pdf_address['city'] != facility.city:
        facility.city = pdf_address['city']
        changed = True
    new_state = normalize_state(pdf_address.get('state', ''))
    if new_state and new_state != facility.state:
        facility.state = new_state
        changed = True
    if pdf_address.get('zip_code') and pdf_address['zip_code'] != facility.zip_code:
        facility.zip_code = pdf_address['zip_code']
        changed = True
    return changed


def _apply_aphis_metadata(facility: PuppyMillFacility, inspections: list[dict]) -> None:
    latest = inspections[0]

    species = set()
    for insp in inspections:
        for field in ('species', 'speciesInspected', 'animalType'):
            val = insp.get(field)
            if val:
                for part in str(val).replace(';', ',').replace('/', ',').split(','):
                    part = part.strip()
                    if part:
                        species.add(part)

    owners = []
    for field in ('legalName', 'customerName', 'ownerName'):
        val = latest.get(field)
        if val and val not in owners:
            owners.append(val)

    if not facility.city and latest.get('city'):
        facility.city = latest.get('city')
    if not facility.state and latest.get('state'):
        facility.state = normalize_state(latest.get('state'))
    if not facility.zip_code:
        facility.zip_code = latest.get('zip') or latest.get('zipCode') or ''

    dog_keywords = ('dog', 'canine', 'puppy', 'puppies')
    species_text = ' '.join(species).lower()
    if species:
        facility.is_dog_facility = any(kw in species_text for kw in dog_keywords)
    if owners:
        facility.owners = owners
    if species:
        facility.dog_breeds = sorted(species)

    facility.processed_report_urls = [
        insp.get('reportLink') for insp in inspections if insp.get('reportLink')
    ]


def _upsert_base_record(data: dict) -> tuple[PuppyMillFacility, bool]:
    license_number = data['license_number']
    profile_url = data.get('usda_profile_url', '')
    facility = PuppyMillFacility.objects.filter(license_number=license_number).first()
    if facility:
        facility.name = data['name']
        facility.dba_name = data.get('dba_name', '')
        facility.license_type = data['license_type']
        facility.license_expiration = data.get('license_expiration')
        if profile_url:
            facility.usda_profile_url = profile_url
        facility.save(update_fields=[
            'name', 'dba_name', 'license_type', 'license_expiration', 'usda_profile_url',
        ])
        return facility, False

    facility = PuppyMillFacility.objects.create(
        license_number=license_number,
        name=data['name'],
        dba_name=data.get('dba_name', ''),
        license_type=data['license_type'],
        city=data.get('city', ''),
        state=normalize_state(data.get('state', '')),
        license_expiration=data.get('license_expiration'),
        usda_profile_url=profile_url,
        owners=[data['name']] if data.get('name') else [],
    )
    return facility, True


def import_usda_facilities() -> dict:
    """Refresh the breeder/dealer list from the USDA spreadsheet."""
    summary = {'created': 0, 'updated': 0, 'errors': 0, 'total_usda': 0}
    try:
        usda_facilities = fetch_usda_licensees()
    except Exception as exc:
        logger.exception('Failed to fetch USDA licensee data')
        summary['error'] = str(exc)
        return summary

    summary['total_usda'] = len(usda_facilities)
    for data in usda_facilities:
        try:
            _, created = _upsert_base_record(data)
            if created:
                summary['created'] += 1
            else:
                summary['updated'] += 1
        except Exception:
            logger.exception('Error importing %s', data.get('license_number'))
            summary['errors'] += 1
    return summary


    return summary


def _sync_facility(facility: PuppyMillFacility) -> str:
    """
    Check one facility for new APHIS reports.
    Only writes to the DB when this is a new site or new report URLs appear.
    Already-stored reports are never updated or re-parsed.
    """
    try:
        inspections = list(search_inspections({'certNumber': facility.license_number}))
    except Exception as exc:
        logger.warning('APHIS check failed for %s: %s', facility.license_number, exc)
        facility.last_checked_at = timezone.now()
        facility.save(update_fields=['last_checked_at'])
        raise

    facility.last_checked_at = timezone.now()
    is_new_site = facility.last_scraped_at is None

    if not inspections:
        facility.save(update_fields=['last_checked_at'])
        if is_new_site and not facility.coordinates_geocoded:
            if _geocode_facility(facility):
                facility.last_scraped_at = timezone.now()
                facility.save()
                return 'initialized'
        return 'skipped'

    new_inspections = find_new_inspections(facility, inspections)

    if not is_new_site and not new_inspections:
        facility.save(update_fields=['last_checked_at'])
        return 'unchanged'

    new_reports = create_new_inspection_reports(facility, new_inspections)
    parse_violations_for_reports(new_reports)

    _apply_aphis_metadata(facility, inspections)

    pdf_url = _newest_inspection_url(new_inspections)
    address_changed = False
    if pdf_url:
        pdf_address = fetch_address_from_report_url(pdf_url)
        if pdf_address:
            address_changed = _apply_pdf_address(facility, pdf_address)

    if is_new_site or address_changed or not facility.coordinates_geocoded:
        _geocode_facility(facility)

    recompute_facility_violation_counts(facility)
    facility.last_scraped_at = timezone.now()
    facility.save()

    return 'initialized' if is_new_site else 'updated'


def _sync_facility_timed(facility: PuppyMillFacility) -> str:
    timeout = sync_state.FACILITY_TIMEOUT_SECONDS
    with ThreadPoolExecutor(max_workers=1) as executor:
        future = executor.submit(_sync_facility, facility)
        try:
            return future.result(timeout=timeout)
        except FuturesTimeout:
            logger.warning(
                'Timed out checking %s (%s) after %ss',
                facility.name, facility.license_number, timeout,
            )
            facility.last_checked_at = timezone.now()
            facility.save(update_fields=['last_checked_at'])
            raise


def run_violation_parsing(batch_size: int | None = None) -> dict:
    """Parse pending inspection report PDFs without running a full facility sync."""
    sync_state.clear_stale_lock()
    if sync_state.get_sync_state().is_running:
        return {'skipped': True, 'reason': 'collection already in progress'}

    lock = sync_state.try_acquire_lock()
    if lock is None:
        return {'skipped': True, 'reason': 'collection already in progress'}

    summary = parse_pending_violations_batch(batch_size=batch_size)
    summary['status'] = 'complete'
    summary['completed_at'] = timezone.now().isoformat()
    sync_state.release_lock(lock, summary)
    return summary


def run_collection(force: bool = False, import_usda: bool = True) -> dict:
    """
    Main collector entry point. By default, refreshes the USDA licensee list first,
    then checks each breeder for new APHIS reports.
    Skips breeders already checked within the last 24 hours unless force=True.
    """
    sync_state.clear_stale_lock()
    if sync_state.get_sync_state().is_running:
        return {'skipped': True, 'reason': 'collection already in progress'}

    if not force and not import_usda and not sync_state.is_sync_due():
        return {
            'skipped': True,
            'reason': 'not due yet (runs every 24 hours)',
            'next_sync_at': sync_state.next_sync_at().isoformat(),
        }

    lock = sync_state.try_acquire_lock()
    if lock is None:
        return {'skipped': True, 'reason': 'collection already in progress'}

    summary = {'checked': 0, 'updated': 0, 'unchanged': 0, 'skipped': 0, 'errors': 0, 'timed_out': 0}
    resume_from = int((sync_state.get_progress() or {}).get('resume_from') or 0)
    run_started_at = timezone.now()

    try:
        if import_usda:
            import_summary = import_usda_facilities()
            summary['usda_created'] = import_summary.get('created', 0)
            summary['usda_updated'] = import_summary.get('updated', 0)
            if import_summary.get('error'):
                summary['status'] = 'error'
                summary['error'] = import_summary['error']
                sync_state.release_lock(lock, summary)
                return summary

        facilities = list(
            PuppyMillFacility.objects.filter(
                Q(is_dog_facility=True) | Q(last_scraped_at__isnull=True)
            ).order_by('id')
        )
        total = len(facilities)
        check_cutoff = timezone.now() - timedelta(hours=sync_state.sync_interval_hours())

        for i, facility in enumerate(facilities, start=1):
            if resume_from and i < resume_from:
                continue
            if not force and facility.last_checked_at and facility.last_checked_at >= check_cutoff:
                continue

            sync_state.set_progress(
                i, total, f'Checking {facility.name} ({i}/{total})…',
            )
            try:
                result = _sync_facility_timed(facility)
                summary['checked'] += 1
                summary[result] = summary.get(result, 0) + 1
                if result == 'updated':
                    time.sleep(sync_state.aphis_delay())
            except FuturesTimeout:
                summary['errors'] += 1
                summary['timed_out'] += 1
            except Exception:
                logger.exception('Error checking %s', facility.license_number)
                summary['errors'] += 1

        sync_state.set_progress(
            total, total, 'Parsing new inspection report violations…',
        )
        summary.update(parse_pending_violations_batch(created_since=run_started_at))

        summary['status'] = 'complete'
        summary['completed_at'] = timezone.now().isoformat()
        sync_state.release_lock(lock, summary)
        logger.info('Dog Watch collection complete: %s', summary)
        return summary
    except Exception as exc:
        summary['status'] = 'error'
        summary['error'] = str(exc)
        sync_state.release_lock(lock, summary)
        raise
