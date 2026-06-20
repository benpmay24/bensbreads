"""Scrape USDA licensee data and enrich with APHIS inspection records."""
import io
import logging
import time
from datetime import datetime

import openpyxl
import requests
from django.utils import timezone

from main.models import PuppyMillFacility
from main.dog_watch.aphis_client import (
    _build_inspection_reports,
    _inspection_report,
    _int_field,
    search_inspections,
)
from main.dog_watch.geocoder import geocode, normalize_state
from main.dog_watch.report_address import fetch_address_from_report_url
from main.dog_watch.news import fetch_news
from main.dog_watch import sync_state

logger = logging.getLogger(__name__)

USDA_LICENSEE_URL = (
    'https://www.aphis.usda.gov/sites/default/files/'
    'list-of-active-licensees-and-registrants.xlsx'
)
TARGET_LICENSE_TYPES = {'Class A - Breeder', 'Class B - Dealer'}
HEADER_ROW_LABEL = 'Registration Type'


def set_progress(phase: str, current: int, total: int, message: str = '') -> None:
    state = sync_state.get_sync_state()
    state.progress = {
        'running': True,
        'phase': phase,
        'current': current,
        'total': total,
        'message': message,
        'updated_at': timezone.now().isoformat(),
    }
    state.save(update_fields=['progress'])


def clear_progress() -> None:
    state = sync_state.get_sync_state()
    state.progress = {'running': False}
    state.save(update_fields=['progress'])


def get_progress() -> dict:
    return sync_state.get_progress()


def _parse_expiration(value) -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    try:
        return datetime.strptime(str(value).strip(), '%m/%d/%Y').date()
    except ValueError:
        return None


def fetch_usda_licensees() -> list[dict]:
    """Download and parse the USDA active licensees spreadsheet."""
    set_progress('download', 0, 1, 'Downloading USDA licensee data...')
    res = requests.get(USDA_LICENSEE_URL, timeout=60)
    res.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(res.content), read_only=True)
    ws = wb[wb.sheetnames[0]]

    facilities = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        if row[1] == HEADER_ROW_LABEL:
            headers = row
            continue
        if not headers or not row[1] or not str(row[1]).startswith('Class '):
            continue
        license_type = row[1]
        if license_type not in TARGET_LICENSE_TYPES:
            continue
        row_data = dict(zip(headers, row))
        license_number = row_data.get('APHIS Registration Number') or row_data.get('APHIS License Number') or ''
        if not license_number:
            continue
        facilities.append({
            'license_number': str(license_number).strip(),
            'license_type': license_type,
            'name': (row_data.get('Account Name') or '').strip(),
            'dba_name': (row_data.get('DBA Name(s)') or '').strip(),
            'city': (row_data.get('Mailing City') or row_data.get('City') or '').strip(),
            'state': (row_data.get('State Abbreviation') or row_data.get('State') or '').strip(),
            'license_expiration': _parse_expiration(row_data.get('Expiration Date')),
            'usda_profile_url': (
                f'https://efile.aphis.usda.gov/PublicSearchTool/s/'
                f'inspection-reports?certNumber={license_number}'
            ),
        })
    return facilities


def _geocode_facility(facility: PuppyMillFacility) -> bool:
    """Update facility coordinates from its address fields. Returns True if geocoded."""
    state = normalize_state(facility.state)
    if not facility.city or not state:
        return False

    lat, lng, geocoded = geocode(
        facility.name,
        facility.city,
        state,
        facility.street_address,
        facility.zip_code,
    )
    if not geocoded or lat is None or lng is None:
        return False

    facility.latitude = lat
    facility.longitude = lng
    facility.coordinates_geocoded = True
    if state != facility.state:
        facility.state = state
    return True


def _known_report_urls(facility: PuppyMillFacility) -> set[str]:
    urls = set(facility.processed_report_urls or [])
    for report in facility.inspection_reports or []:
        url = report.get('url')
        if url:
            urls.add(url)
    return urls


def _apply_pdf_address(facility: PuppyMillFacility, pdf_address: dict) -> bool:
    """Apply address from a report PDF. Returns True if any field changed."""
    changed = False
    if pdf_address.get('street_address') and pdf_address['street_address'] != facility.street_address:
        facility.street_address = pdf_address['street_address']
        changed = True
    if pdf_address.get('city') and pdf_address['city'] != facility.city:
        facility.city = pdf_address['city']
        changed = True
    new_state = normalize_state(pdf_address.get('state', ''))
    if new_state and new_state != facility.state:
        facility.state = new_state
        changed = True
    if pdf_address.get('zip_code') and pdf_address['zip_code'] != facility.zip_code:
        facility.zip_code = pdf_address['zip_code']
        changed = True
    return changed


def _apply_aphis_metadata(facility: PuppyMillFacility, inspections: list[dict]) -> None:
    """Update violation totals and display reports from APHIS inspection metadata."""
    latest = inspections[0]
    direct = sum(_int_field(i, 'direct', 'directViolations') for i in inspections)
    critical = sum(_int_field(i, 'critical', 'criticalViolations') for i in inspections)
    non_critical = sum(_int_field(i, 'nonCritical', 'nonCriticalViolations') for i in inspections)

    species = set()
    for insp in inspections:
        for field in ('species', 'speciesInspected', 'animalType'):
            val = insp.get(field)
            if val:
                for part in str(val).replace(';', ',').replace('/', ',').split(','):
                    part = part.strip()
                    if part:
                        species.add(part)

    owners = []
    for field in ('legalName', 'customerName', 'ownerName'):
        val = latest.get(field)
        if val and val not in owners:
            owners.append(val)

    if not facility.city and latest.get('city'):
        facility.city = latest.get('city')
    if not facility.state and latest.get('state'):
        facility.state = normalize_state(latest.get('state'))
    if not facility.zip_code:
        facility.zip_code = latest.get('zip') or latest.get('zipCode') or ''

    dog_keywords = ('dog', 'canine', 'puppy', 'puppies')
    species_text = ' '.join(species).lower()
    if species:
        facility.is_dog_facility = any(kw in species_text for kw in dog_keywords)
    if owners:
        facility.owners = owners
    if species:
        facility.dog_breeds = sorted(species)

    facility.violation_count = direct + critical + non_critical
    facility.direct_violations = direct
    facility.critical_violations = critical
    facility.inspection_reports = _build_inspection_reports(inspections)
    facility.processed_report_urls = [
        insp.get('reportLink') for insp in inspections if insp.get('reportLink')
    ]


def _newest_report_url(reports: list[dict]) -> str:
    if not reports:
        return ''
    return max(reports, key=lambda r: r.get('date') or '').get('url') or ''


def _upsert_base_record(data: dict) -> tuple[PuppyMillFacility, bool]:
    license_number = data['license_number']
    profile_url = data.get('usda_profile_url', '')
    facility = PuppyMillFacility.objects.filter(license_number=license_number).first()
    if facility:
        facility.name = data['name']
        facility.dba_name = data.get('dba_name', '')
        facility.license_type = data['license_type']
        facility.license_expiration = data.get('license_expiration')
        if profile_url:
            facility.usda_profile_url = profile_url
        facility.save(update_fields=[
            'name', 'dba_name', 'license_type', 'license_expiration', 'usda_profile_url',
        ])
        return facility, False

    facility = PuppyMillFacility.objects.create(
        license_number=license_number,
        name=data['name'],
        dba_name=data.get('dba_name', ''),
        license_type=data['license_type'],
        city=data.get('city', ''),
        state=normalize_state(data.get('state', '')),
        license_expiration=data.get('license_expiration'),
        usda_profile_url=profile_url,
        owners=[data['name']] if data.get('name') else [],
    )
    return facility, True


def import_usda_facilities() -> dict:
    """Import or update all facilities from the USDA spreadsheet."""
    summary = {'created': 0, 'updated': 0, 'errors': 0, 'total_usda': 0}

    try:
        usda_facilities = fetch_usda_licensees()
    except Exception as exc:
        logger.exception('Failed to fetch USDA licensee data')
        summary['error'] = str(exc)
        return summary

    summary['total_usda'] = len(usda_facilities)
    total = len(usda_facilities)

    for i, data in enumerate(usda_facilities, start=1):
        if i % 50 == 0 or i == total:
            set_progress('import', i, total, f'Importing USDA records ({i}/{total})...')
        try:
            _, created = _upsert_base_record(data)
            if created:
                summary['created'] += 1
            else:
                summary['updated'] += 1
        except Exception:
            logger.exception('Error importing %s', data.get('license_number'))
            summary['errors'] += 1

    return summary


def _sync_facility(
    facility: PuppyMillFacility,
    fetch_news_articles: bool = False,
    progress_index: int | None = None,
    progress_total: int | None = None,
) -> str:
    """
    Check APHIS for new inspection reports and update only when needed.

    Returns: 'unchanged', 'initialized', 'updated', or 'skipped'.
    """
    try:
        inspections = list(search_inspections({'certNumber': facility.license_number}))
    except Exception as exc:
        logger.warning('APHIS check failed for %s: %s', facility.license_number, exc)
        facility.last_checked_at = timezone.now()
        facility.save(update_fields=['last_checked_at'])
        raise

    facility.last_checked_at = timezone.now()
    is_initial = facility.last_scraped_at is None

    if not inspections:
        facility.save(update_fields=['last_checked_at'])
        if is_initial and not facility.coordinates_geocoded:
            if _geocode_facility(facility):
                facility.last_scraped_at = timezone.now()
                facility.save()
                return 'initialized'
        return 'skipped'

    api_reports = [_inspection_report(insp) for insp in inspections]
    known_urls = _known_report_urls(facility)

    if is_initial:
        new_reports = [r for r in api_reports if r.get('url') and r['url'] not in known_urls]
    else:
        latest_known_date = max(
            (r.get('date') or '' for r in (facility.inspection_reports or [])),
            default='',
        )
        new_reports = [
            r for r in api_reports
            if r.get('url')
            and r['url'] not in known_urls
            and (r.get('date') or '') > latest_known_date
        ]

    if not is_initial and not new_reports:
        facility.processed_report_urls = [
            insp.get('reportLink') for insp in inspections if insp.get('reportLink')
        ]
        facility.save(update_fields=['last_checked_at', 'processed_report_urls'])
        return 'unchanged'

    _apply_aphis_metadata(facility, inspections)

    if is_initial:
        pdf_url = _newest_report_url(api_reports)
    else:
        pdf_url = _newest_report_url(new_reports)

    address_changed = False
    if pdf_url:
        if progress_index is not None and progress_total is not None:
            set_progress(
                'check',
                progress_index,
                progress_total,
                f'Reading report PDF for {facility.name} ({progress_index}/{progress_total})...',
            )
        pdf_address = fetch_address_from_report_url(pdf_url)
        if pdf_address:
            address_changed = _apply_pdf_address(facility, pdf_address)

    if is_initial or address_changed or not facility.coordinates_geocoded:
        _geocode_facility(facility)

    if fetch_news_articles and facility.violation_count > 0 and not facility.news_articles:
        facility.news_articles = fetch_news(facility.name, facility.city, facility.state)

    facility.last_scraped_at = timezone.now()
    facility.save()

    if is_initial:
        return 'initialized'
    return 'updated'


def check_all_facilities(fetch_news_articles: bool = False) -> dict:
    """Check every facility for new APHIS reports; update only when reports are new."""
    summary = {
        'checked': 0,
        'unchanged': 0,
        'initialized': 0,
        'updated': 0,
        'skipped': 0,
        'errors': 0,
    }
    facilities = list(
        PuppyMillFacility.objects.filter(is_dog_facility=True).order_by('id')
    )
    total = len(facilities)
    resume_from = int((get_progress() or {}).get('resume_from') or 0)
    if resume_from > 1:
        set_progress(
            'check',
            resume_from,
            total,
            f'Resuming from facility {resume_from}/{total}...',
        )

    for i, facility in enumerate(facilities, start=1):
        if resume_from and i < resume_from:
            continue
        set_progress(
            'check',
            i,
            total,
            f'Checking {facility.name} ({i}/{total})...',
        )
        try:
            result = _sync_facility(
                facility,
                fetch_news_articles=fetch_news_articles,
                progress_index=i,
                progress_total=total,
            )
            summary['checked'] += 1
            summary[result] = summary.get(result, 0) + 1
            if result in ('initialized', 'updated'):
                time.sleep(sync_state.aphis_delay())
        except Exception:
            logger.exception('Error checking %s', facility.license_number)
            summary['errors'] += 1

    return summary


def geocode_pending_facilities() -> dict:
    """Geocode facilities that still lack verified map coordinates."""
    summary = {'geocoded': 0, 'skipped': 0, 'errors': 0}
    facilities = list(
        PuppyMillFacility.objects.filter(
            is_dog_facility=True,
            coordinates_geocoded=False,
        )
        .exclude(city='')
        .exclude(state='')
        .order_by('id')
    )
    total = len(facilities)

    for i, facility in enumerate(facilities, start=1):
        set_progress('geocode', i, total, f'Locating addresses ({i}/{total})...')
        try:
            if _geocode_facility(facility):
                facility.save(update_fields=['latitude', 'longitude', 'coordinates_geocoded', 'state'])
                summary['geocoded'] += 1
            else:
                summary['skipped'] += 1
        except Exception:
            logger.exception('Error geocoding %s', facility.license_number)
            summary['errors'] += 1

    return summary


def run_full_sync(
    force: bool = False,
    fetch_news_articles: bool = False,
) -> dict:
    """
    Sync: refresh USDA licensee list, check each facility for new reports only.
    Runs when due (every 24 hours) or when forced via Update Now.
    """
    sync_state.clear_stale_lock()

    if get_progress().get('running'):
        return {'skipped': True, 'reason': 'sync already in progress'}

    if not force and not sync_state.is_sync_due():
        return {
            'skipped': True,
            'reason': 'sync not due yet',
            'next_sync_at': sync_state.next_sync_at().isoformat(),
        }

    lock = sync_state.try_acquire_lock()
    if lock is None:
        return {'skipped': True, 'reason': 'sync already running on another worker'}

    summary = {'skipped': False, 'status': 'complete'}

    try:
        set_progress('scheduled', 0, 0, 'Checking for new inspection reports...')

        import_summary = import_usda_facilities()
        summary.update({
            'created': import_summary.get('created', 0),
            'usda_updated': import_summary.get('updated', 0),
            'total_usda': import_summary.get('total_usda', 0),
        })
        if import_summary.get('error'):
            summary['error'] = import_summary['error']
            summary['status'] = 'error'
            sync_state.release_lock(lock, summary)
            return summary

        check_summary = check_all_facilities(fetch_news_articles=fetch_news_articles)
        summary['checked'] = check_summary.get('checked', 0)
        summary['unchanged'] = check_summary.get('unchanged', 0)
        summary['initialized'] = check_summary.get('initialized', 0)
        summary['new_reports'] = check_summary.get('updated', 0)
        summary['skipped'] = check_summary.get('skipped', 0)
        summary['errors'] = check_summary.get('errors', 0)

        geocode_summary = geocode_pending_facilities()
        summary['geocoded'] = geocode_summary.get('geocoded', 0)
        summary['errors'] += geocode_summary.get('errors', 0)

        summary['mapped_count'] = PuppyMillFacility.objects.filter(
            is_dog_facility=True,
            coordinates_geocoded=True,
        ).count()
        summary['total_facilities'] = PuppyMillFacility.objects.filter(
            is_dog_facility=True,
        ).count()
        summary['completed_at'] = timezone.now().isoformat()
        summary['next_sync_at'] = sync_state.next_sync_at().isoformat()

        logger.info('Dog Watch sync complete: %s', summary)
        sync_state.release_lock(lock, summary)
        return summary
    except Exception as exc:
        summary['status'] = 'error'
        summary['error'] = str(exc)
        sync_state.release_lock(lock, summary)
        raise


run_scheduled_sync = run_full_sync


def scrape_puppy_mills(
    enrich: bool = True,
    fetch_news_articles: bool = False,
    force_import: bool = False,
) -> dict:
    """CLI entry point: import USDA data and optionally run a sync."""
    summary = {'created': 0, 'updated': 0, 'errors': 0, 'total_usda': 0}

    if force_import or PuppyMillFacility.objects.count() == 0:
        import_summary = import_usda_facilities()
        summary.update(import_summary)
        if import_summary.get('error'):
            return summary

    if enrich:
        check_summary = check_all_facilities(fetch_news_articles=fetch_news_articles)
        summary['checked'] = check_summary.get('checked', 0)
        summary['new_reports'] = check_summary.get('updated', 0)
        summary['errors'] += check_summary.get('errors', 0)
        geocode_summary = geocode_pending_facilities()
        summary['geocoded'] = geocode_summary.get('geocoded', 0)
        summary['errors'] += geocode_summary.get('errors', 0)

    return summary
